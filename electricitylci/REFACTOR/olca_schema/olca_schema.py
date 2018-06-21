import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import math
import os
import sys
import scipy
from scipy.stats import t
from sympy import *
import pickle
import time
import warnings




warnings.filterwarnings("ignore")

#Reading the name of flow file
names = pd.read_excel('eLCI_data.xlsx', sheet_name='generation_elecname')

#26 egrid regions in an array
ar_reg = ["AKMS","AZNM","CAMX","ERCT","FRCC","HIMS","HIOA","MROE","MROW","NEWE","NWPP","NYCW","NYLI","NYUP","RFCE","RFCM","RFCW","RMPA","SPNO","SPSO","SRMV","SRMW","SRSO","SRTV","SRVC"]



#Reading the elementary flow list file WILL COME FROM FEDERAL LCA ELEMENTARY FLOW DATABASE. CALL FUNCTION FOR THAT> 
elem_flow_map_orig = pd.read_csv('USEEIO_FlowMapping.csv', header = 0)


#Reading the facility file from STEWI (will be replaaced with a call to STEWI )
egrid1_orig = pd.read_csv("eGRID_2014_1.csv", header=0, error_bad_lines=False)

#Reading the Stewi combo file. (double reading of same file needs to be corrected) Replaced by call to STEWI COMBO
stewi_combo_orig = pd.read_csv("egrid2014_trircrainfoneiegrid.csv", header=0, error_bad_lines=False)  


#Reading the flow by facility file coming from STEWI Combo. 
egrid2_2_orig = pd.read_csv("egrid2014_trircrainfoneiegrid.csv", header=0, error_bad_lines=False)

#Reading eia from pickle file as main file is too slow. 
#eia_orig = pd.read_csv('EIA923_Schedules_2_3_4_5_M_12_2015_Final_Revision.csv',skipinitialspace = True)
eia_orig = pickle.load(open( "save.p", "rb" ))


#Reading the fuel name file
fuel_name = pd.read_excel('eLCI_data.xlsx', sheet_name='fuelname')




#This is a small function to check for year and decide what years need to be downloaded. Could be simpler because
#if we specify the eGRID year as a parameter,an specify the years for the inventories, then we dont need this function. This will make it faster
def year_check():
    
    global odd_year;
    global odd_database;
    global stewi_combo_orig;
    global year;
    egrid_years = [2014,2016];
    egrid_time = pd.DataFrame() 
    egrid_time[['Year','Source']]= stewi_combo_orig[['Year','Source']]
    
    egrid_time1 = egrid_time.groupby(by = ['Source'])['Year'].mean()
    egrid_time1 = egrid_time1.reset_index()
    
    for row in egrid_time1[['Year']].itertuples():
        if(row[1] != 2014 and row[1] != 2016):
             odd_year = row[1]
             odd_database = row[0]
             
        else:
             year = row[1]

#This is the name changer function that helps to replace the chemical names with the names from the elementary flows list. GLOBALS - Change names to something understandable
def name_changer(inventory):
        

        elem_flow_map = elem_flow_map_orig[['OriginalName', 'OriginalCategory','Source','NewName']]
        #Category name has starting letter capital. Reducing to lower case. 
        elem_flow_map['OriginalCategory'] = elem_flow_map['OriginalCategory'].str.lower()
        elem_flow_map = elem_flow_map.reset_index()
        
        #Merging old name inventory with new name inventory
        new_data = pd.merge(inventory, elem_flow_map,left_on=['FlowName','Compartment','Source'], right_on=['OriginalName', 'OriginalCategory','Source'],how='left')
        
        #Replacing the blanks that are the missing names matches with wrong word
        new_data['NewName'] = new_data.loc[:,['NewName']].replace(np.nan, 'wrong', regex=True)

        #Only replacing those names which do not have the wrong keyword. The ones with the wrong keyword are kept as the original names. 
        new_data['FlowName'] = np.where(new_data['NewName'] != 'wrong',new_data['NewName'],new_data['FlowName'])
        #Dropping unnecessary columns
        new_data1 = new_data.drop(['OriginalName', 'OriginalCategory','NewName','index'],axis = 1)
        
              
        return new_data1

#This is the function for processing the eGRID dataset. 
def egrid_func(a,b):


        global egrid2_2_orig;
        
        #First sending the inventory to the name changer for changing all names. 
        egrid2 = name_changer(egrid2_2_orig)
        
        #Extracting out only the eGRID emissions from the file coming from STEWI combo. 
        egrid2 = egrid2[egrid2['Source'] == 'eGRID']
        
        #Removing all duplicates
        egrid2 = egrid2.drop_duplicates(keep = 'first')
        
        #pivoting the database from a list format to a table format
        egrid3 = pd.pivot_table(egrid2,index = 'FacilityID' ,columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1_orig.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        
        #Dropping all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
        
        
        #Dropping less than a% and more than b% efficiency.
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        
        
       
        #shifitng the column Heat input to the front of the dataframe
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        #shifitng the column Net generation to the front of the dataframe and renaming it. 
        col = list(egrid_new1['Net generation'])         
        egrid_new1['NetGen'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net generation'])       
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        egrid_new1['NetGen'] = egrid_new1['NetGen']*0.277778/1000
        #egrid_new1.to_csv('egrid_revised.csv')
        return egrid_new1


#THis is for processing the EIA 923 data So this can be reduced to one function call for efficiencies.
def egrid_func2(a,b):
        
        #Extracing out few columns necessary
        eia_w = eia_orig[['Plant Id','Plant Name','Plant State','Total Fuel ConsumptionMMBtu','Net Generation']]
        
       
        #Grouping similar facilities together.         
        eia_w1 = eia_w.groupby(['Plant Id','Plant Name','Plant State'],as_index = False)['Total Fuel ConsumptionMMBtu','Net Generation'].sum()
        
        
        #I used egrid as variable names because I copy pasted the code from the previous function. Dont confude the variable name with the actual database being processed.
        
        egrid = eia_w1;
        egrid[['Total Fuel ConsumptionMMBtu','Net Generation']] = egrid[['Total Fuel ConsumptionMMBtu','Net Generation']].apply(pd.to_numeric,errors = 'coerce')
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net Generation'].values*100/(egrid['Total Fuel ConsumptionMMBtu'].values*0.29307)
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        #Replacing all the facilities with no net generation or Heat input reported
        egrid = egrid.dropna(subset = ['Efficiency'])
        
        #Dropping less than a% and more than b%        
        egrid_new2 = egrid[egrid['Efficiency'].values > float(a)]        
        egrid_new1 = egrid_new2[egrid_new2['Efficiency'] < float(b)]
        
        #shifitng the column Heat input to the front of the dataframe       
        col = list(egrid_new1['Total Fuel ConsumptionMMBtu']*0.29307)        
        egrid_new1['HeatInput'] = col
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        
        #shifitng the column Net generation to the front of the dataframe and renaming it. 
        col = list(egrid_new1['Net Generation'])         
        egrid_new1['NetGen_new'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(columns = ['Net Generation'])       
        egrid_new2 = egrid_new1[['NetGen_new','Plant Id']]      

        
        return egrid_new2


#TRoy weight based method to compute emissions factors.  Rename calculation of emission factors. descriptive of the varaible.  
def compilation(db):
        #Troy Method
        #Creating copy of database by substitution the NA emissions with zero
        db1 = db.fillna(value = 0)
        
        #Removing all rows here emissions are not reported for second dataframe
        db2 = db.dropna()
        
        
        #keeping the unreported emissions and facilities in separate database

        #This check is to make sure that the second database is not empt after droppins all NA. if empty, then we only use first database.  
        if db2.empty == True:
            ef1 = np.sum(db1.iloc[:,1])/np.sum(db1.iloc[:,0])
            return ef1
    
        ef1 = np.sum(db1.iloc[:,1])/np.sum(db1.iloc[:,0])
        
        ef2 = np.sum(db2.iloc[:,1])/np.sum(db2.iloc[:,0])
        
        #weight formula.
        weight = np.sum(db2.iloc[:,0])/np.sum(db1.iloc[:,0])
        final_ef = ef2*weight + (1-weight)*ef1
        
        return final_ef




def tri_func(data,a,b):
        
        #Reading tri database                
        global stewi_combo_orig;
        
        #Droppping duplicates in STEWI
        stewi_combo1_1 = stewi_combo_orig.drop_duplicates(keep = 'first')
        
        #Sending to name changing function. 
        stewi_combo = name_changer(stewi_combo1_1)
        
        #Extracting the columns from eGRID. 
        stewi_combo = stewi_combo[['eGRID_ID','FlowName','Compartment','FlowAmount','Source','ReliabilityScore','Year']]
        
        #This is done to combine Stack and Fugutive emissions in NEI and the different Stream information in the TRI to just one flow by adding them up. 
        stewi_combo2 = stewi_combo.groupby(['eGRID_ID','FlowName','Compartment','Source','ReliabilityScore','Year'])['FlowAmount'].sum()
        stewi_combo2 = stewi_combo2.reset_index()
        
        #Taking all other flows apart from eGRID.
        stewi_combo2 = stewi_combo2[stewi_combo2['Source']!='eGRID']
        
       
        #stewi_combo2 = stewi_combo2.dropna(axis = 0, how = 'any')
        #stewi_combo reshaping pivot not working so using pivot table
        stewi_combo3 = pd.pivot_table(stewi_combo2,index = ['eGRID_ID','Compartment','Source','ReliabilityScore','Year'], columns = 'FlowName', values = 'FlowAmount')
        
        stewi_combo3 = stewi_combo3.reset_index()

        
        #This is the eGRID database which is used for merging with the pther databases such that only the plants that were chosen to have within limit efficiencies are also chosen from the other databases. 
        data = data[['NetGen','FacilityID']]
        
             
        #Merging egrid and other databases. 
        database = data.merge(stewi_combo3,left_on ='FacilityID', right_on = 'eGRID_ID')
        
        
        global odd_year;
        
        
        #Calling the EIA for replacing those emissiosn generation that are not there in the eGRID YEAR and needs to be obtained from EIA. 
        eia = egrid_func2(a,b);
        
        
        database = eia.merge(database,left_on = 'Plant Id',right_on = 'eGRID_ID')
        
        #need to figure out how to know 2015 is the odd one out, 
        #Reaplcing the odd year Net generations with the EIA net generations. 
        database['NetGen']= np.where(database['Year'] == odd_year, database['NetGen_new'],database['NetGen'])
        
        #Dropping unnecessary columns
        database = database.drop(columns = ['FacilityID','eGRID_ID','Plant Id','Year','NetGen_new'])
        
        return database  
                   
         
# A major refactor template writer py file separate from everything. Starting to store the data in some dataframes.     
#have the dataframe in same types of table. 
        #easier to write JSONLD

#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br


#This is the function for calculating log normal distribution parameters. 
def uncertainty(db):
        #Troy Method
        #Creating copy of database by substitution the NA emissions with zero
        db1 = db.fillna(value = 0)
        
        #Removing all rows here emissions are not reported for second dataframe
        db2 = db.dropna()
        frames = [db1,db2]
        #Here we doubled up the database by combining two databases together
        data = pd.concat(frames,axis = 0)
        
        mean = np.mean(data.iloc[:,1])
        l,b = data.shape
        sd = np.std(data.iloc[:,1])/np.sqrt(l)
        mean_gen = np.mean(data.iloc[:,0])
        #obtaining the emissions factor from the weight based method
        ef = compilation(db)
        
        #Endpoints of the range that contains alpha percent of the distribution
        pi1,pi2 = t.interval(alpha = 0.90,df = l-2, loc = mean, scale = sd)
        #Converting prediction interval to emission factors
        pi2 = pi2/mean_gen
        pi1 = pi1/mean_gen
        pi3 = (pi2-ef)/ef;
        x = var('x')
        
        
    
        
        if math.isnan(pi3) == True:
            return None,None;
        
        elif math.isnan(pi3) == False:
            
            #This method will not work with the interval limits are more than 280% of the mean. 
            if pi3 < 2.8:
              sd1,sd2 = solve(0.5*x*x -(1.16308*np.sqrt(2))*x + (np.log(1+pi3)),x)

            else:#This is a wrong mathematical statement. However, we have to use it if something fails. 
              sd1,sd2 = solve(0.5*x*x -(1.36*np.sqrt(2))*x + (np.log(1+pi3)),x)
              print(pi3)
            
            #always choose lower standard deviation from solving the square root equation. 
            if sd1 < sd2:
               log_mean = np.log(ef)-0.5*(sd1**2)
               return round(log_mean,2),round(sd1,2)
            else:
               log_mean = np.log(ef)-0.5*(sd2**2)
               return round(log_mean,2),round(sd2,2)



def initial_dictionary_creation():
        olca_schema_dict = pickle.load(open( "olca_schema_dict.pk", "rb" ))
        
        #print(list(olca_schema_dict.keys()))
        
        #['Actor', 'AllocationFactor', 'CalculationSetup', 'CategorizedEntity', 'Category', 'Entity', 'Exchange', 'Flow', 'FlowProperty', 'FlowPropertyFactor', 'FlowRef', 'FlowResult', 'ImpactCategory', 'ImpactCategoryRef', 'ImpactFactor', 'ImpactMethod', 'ImpactResult', 'Location', 'Parameter', 'ParameterRedef', 'Process', 'ProcessDocumentation', 'ProcessLink', 'ProcessRef', 'ProductSystem', 'Ref', 'RootEntity', 'SimpleResult', 'SocialIndicator', 'Source', 'Uncertainty', 'Unit', 'UnitGroup']
        
        ar = ['Process','ProcessDocumentation','Exchange','Uncertainty','Flow']
        
        cols_for_exchanges_df = list(olca_schema_dict['Process'].keys())
        d = pd.DataFrame(columns=cols_for_exchanges_df)
        
       
        
#        for i in ar:
#            cols_for_exchanges_df = list(olca_schema_dict[i].keys())
#            d[i] = pd.DataFrame(columns=cols_for_exchanges_df)
        
#        del d['key']
        
        return d;


    

    

        
        



def olca_schema_generator(l_limit,u_limit,Reg):
        
        global region;
        region = Reg;
        
        year_check()
        
        database = egrid_func(l_limit,u_limit)
        
        #Extracing out only part of the database that belongs to a specific eGRID region. 
        database = database[database['eGRID subregion acronym'] == Reg]
        
        
        #Dictionary of lists 
        d_list = ['eGRID','NEI','TRI','RCRAInfo']
                            
        global fuel_name;
        global names;
        global year;
        global odd_year;
        global odd_database;
        global database_f1;
        global fuelheat;
        global fuelname;
            
        for row in fuel_name.itertuples():
        #assuming the starting row to fill in the emissions is the 5th row Blank
            
            global blank_row;    
            blank_row = 6;
            index = 2;
    
                       
            #Reading complete fuel name and heat content information       
            fuelname = row[2]
            fuelheat = float(row[3])
            

            v= [];
            v1= [];
            
            
            #THis code block is used for finding the NERC regions and states.          
            for roww in database.itertuples():
                if row[1] == roww[9] or row[1] == roww[10]:
                    v1.append(roww[4])
                    v.append(roww[11])
                    break;
            v2 = list(set(v1)) 
            str1 = ','.join(v2)
            v2 = list(set(v))
            str2 = ','.join(v2)
           

            
                 

            
            
            #croppping the database according to the current fuel being considered
            database_f1 = database[database['Plant primary coal/oil/gas/ other fossil fuel category'] == row[1]]
            if database_f1.empty == True:
                  database_f1 = database[database['Plant primary fuel'] == row[1]]  
            
                
            if database_f1.empty != True:       
            
                    #This part is used for writing the input fuel flow informationn
                    if database_f1['HeatInput'].mean() != 0 and fuelheat != 0:
                        
                        ra1 = exchange_table_creation_input(database_f1);
                                  
                    #This part is used for filling up the emission information from the different databases. 
                    def flowwriter(database_f1,y):
    
                                             
                        for i in database_f1.iteritems():
                          
                          #Only writng the emissions. NOt any other flows or columns in the template files.   
                          if str(i[0]) != 'NetGen' and str(i[0]) != 'ReliabilityScore': 
                            database_f2 = database_f1[['NetGen',i[0]]] 
                                
                            if(compilation(database_f2) != 0 and compilation(database_f2)!= None):
                                
                                
                                
                                ra = exchange_table_creation_output(database_f2,y)
                                return ra;
                                
                                                   
                    
                    
                    
                    
                    for x in d_list:
                        
                        if x == 'eGRID': 
                            database_f3 = database_f1[['NetGen','Carbon dioxide', 'Nitrous oxide', 'Nitrogen oxides', 'Sulfur dioxide', 'Methane']]
                            ra2 = flowwriter(database_f3,x)
                            
                        elif x != 'eGRID':  
                            database_f3 = tri_func(database_f1,l_limit,u_limit);
                            
                            #This is for extracing only the database being considered fro the d list names. 
                            if x == 'TRI':                     
                                database_f3 = database_f3[database_f3['Source']=='TRI']
                           
                            elif x == 'NEI':
                                database_f3 = database_f3[database_f3['Source']=='NEI']
                           
                            elif x == 'RCRAInfo':
                                database_f3 = database_f3[database_f3['Source']=='RCRAInfo']
                            
                            #CHecking if its not empty and differentiating intp the different Compartments that are possible, air , water soil and wastes. 
                            if database_f3.empty != True:
                                                      
                                #water
                                d1 = database_f3[database_f3['Compartment']=='air']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  
                                  ra2 = flowwriter(d1,x)                           
                                
                                
                                #water
                                d1 = database_f3[database_f3['Compartment']=='water']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  ra2 = flowwriter(d1,x)  
                                
                                #soil
                                d1 = database_f3[database_f3['Compartment']=='soil']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  ra2 = flowwriter(d1,x)  
                                
                                                        
                                #waste
                                d1 = database_f3[database_f3['Compartment']=='waste']
                                d1 = d1.drop(columns = ['Compartment','Source'])
                                
                                if d1.empty != True:
                                  ra2 = flowwriter(d1,x)  
            
                             
                        
                            
                            
                    #Writing final file. 
                    
                    global process
                    data_dir = os.path.dirname(os.path.realpath(__file__))+"\\results\\"
                    final = process_table_creation(ra1,ra2)
                    del final['']
                    process[fuelname+'_'+Reg] = final;
                    print(fuelname+'_'+Reg+' File written Successfully')
                    global t0;
                    t1 = time.time()
                    print(t1-t0)
                    t0 = t1;




def process_table_creation(inp,out):
    
    
    global region;
    
                              
    ar = {'':''}
    
    ar['allocationFactors']=''
    ar['defaultAllocationMethod']=''
    ar['exchanges']=[inp,out]
    ar['location']=region
    ar['parameters']=''
    ar['processDocumentation']=process_doc_creation();
    ar['processType']=''
    
    return ar;
    
    

def process_doc_creation():
    
    global year;
    
    

    ar = {'':''}
    ar['timeDescription']=''
    ar['validUntil']='12/31'+str(year)
    ar['validFrom']='1/1/'+str(year)
    ar['technologyDescription']=''
    ar['dataCollectionDescription']=''
    ar['completenessDescription']=''
    ar['dataSelectionDescription']=''
    ar['reviewDetails']=''
    ar['dataTreatmentDescription']=''
    ar['inventoryMethodDescription']=''
    ar['modelingConstantsDescription']=''
    ar['reviewer']='Wes Ingwersen'
    ar['samplingDescription']=''
    ar['sources']=''
    ar['restrictionsDescription']=''
    ar['copyright']=''
    ar['creationDate']=''
    ar['dataDocumentor']='Wes Ingwersen'
    ar['dataGenerator']='Tapajyoti Ghosh'
    ar['dataSetOwner']=''
    ar['intendedApplication']=''
    ar['projectDescription']=''
    ar['publication']=''
    ar['geographyDescription']=''
    
    return ar;




def exchange_table_creation_input(data):
    
    global fuelname;
    global year;
    global database_f1;
    global fuelheat; 
    

    ar = {'':''}
    
    ar['internalId']=''
    ar['avoidedProduct']='False'
    ar['flow']=flow_table_creation(fuelname)
    ar['flowProperty']=''
    ar['input']='True'
    ar['quantitativeReference']='True'
    ar['baseUncertainty']=''
    ar['provider']=''
    ar['amount']=(np.sum(database_f1['HeatInput'])/np.sum(database_f1['NetGen']))/fuelheat;
    ar['amountFormula']=''
    ar['unit']='kg'
    ar['pedigreeUncertainty']=''
    ar['uncertainty']=uncertainty_table_creation(data)
    ar['comment']='eGRID '+str(year);
    
    return ar;
 
def exchange_table_creation_output(data,y):
    
    global d;
    global odd_year;

    ar = {'':''}
    
    ar['internalId']=''
    ar['avoidedProduct']='False'
    
    ar['flow']=flow_table_creation(data.columns[1])
    ar['flowProperty']=''
    ar['input']='False'
    ar['quantitativeReference']=''
    ar['baseUncertainty']=''
    ar['provider']=''
    ar['amount']=compilation(data)
    ar['amountFormula']=''
    ar['unit']='kg'
    ar['pedigreeUncertainty']=''
    ar['uncertainty']=uncertainty_table_creation(data)
    
    if y == odd_database:
         
         ar['comment'] = str(y)+' '+str(odd_year);
         
    else: 
                         
         ar['comment'] = str(y)+' '+str(year);

    
    return ar;    

        
        
def uncertainty_table_creation(data):
    
    global fuelheat;
    
    
    ar = {'':''}
    
    if data.columns[1] == 'HeatInput':
    
            temp_data = data[['NetGen','HeatInput']]
            #uncertianty calculations only if database length is more than 3
            l,b = temp_data.shape
            if l > 3:
               u,s = uncertainty(temp_data)
               if str(fuelheat)!='nan':
                  ar['geomMean'] = str(round(math.exp(u),3)/fuelheat);
                  ar['geomSd']=str(round(math.exp(s),3)); 
               else:
                  ar['geomMean'] = str(round(math.exp(u),3)); 
                  ar['geomSd']=str(round(math.exp(s),3)); 
    
    else:
    
            #uncertianty calculations
                    l,b = data.shape
                    if l > 3:
                       u,s = (uncertainty(data))
                       ar['geomMean'] = str(round(math.exp(u),3)); 
                       ar['geomSd']=str(round(math.exp(s),3)); 

    
    
    ar['distributionType']='Logarithmic Normal Distribution'
    ar['mean']=''
    ar['meanFormula']=''
    
    ar['geomMeanFormula']=''
    ar['minimum']=data.iloc[:,1].min();
    ar['minimumFormula']=''
    ar['sd']=''
    ar['sdFormula']=''
    
    ar['geomSdFormula']=''
    ar['mode']=''
    ar['modeFormula']=''
    ar['maximum']=data.iloc[:,1].max();
    ar['maximumFormula']='';
    
    return ar;


def flow_table_creation(fl):
    
    global region;
    
                    
    ar = {'':''}
    ar['flowType']='PRODUCT_FLOW'
    ar['cas']=str(fl)
    ar['formula']=''
    ar['flowProperties']=''
    ar['location']=str(region)
    
    return ar











 
#generator(10,100,'AZNM')

#Used for speeding up the calculation using parallel processing.
import concurrent.futures



t0 = time.time()
def function(x):
    global process
    process = {'':''}      
    initial_dictionary_creation()
    olca_schema_generator(10,100,x)


def write_json_file():
  global process;
  import json
  with open('data.json', 'w') as f:
     json.dump(process['biomass_AZNM'], f, ensure_ascii=False)
  
'''
ar_reg = ['AZNM']
def main():
     
     if __name__ == '__main__':     
       with concurrent.futures.ProcessPoolExecutor() as executor:
        zip(ar, executor.map(function, ar_reg))

from multiprocessing import Process

Pros = []

def main():
  for i in range(0,2):
     year_check()
     if __name__ == '__main__':
         p = Process(target=function, args=(ar_reg[i],))
         p.start()
         Pros.append(p)
         
  for p in Pros:
     p.join()      

'''

  # block until all the threads finish (i.e. block until all function_x calls finish)    
  #for th in Pros:
  #   th.join()

write_json_file()

#function('AZNM')




    