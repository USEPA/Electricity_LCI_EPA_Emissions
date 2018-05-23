#Writes out generation processes. Not currently working because relies on inventory from old sources

import pandas as pd
from copy import copy, deepcopy
import openpyxl
import numpy as np
import os
import sys

def egrid_func(a,b):

        #Reading the facility file
        egrid1 = pd.read_csv("egrid_2016_1.csv", header=0, error_bad_lines=False)
        
        
        #Reading the flow by facility file
        egrid2 = pd.read_csv("egrid_2016.csv", header=0, error_bad_lines=False)
        
        
        egrid3 = egrid2.pivot(index = 'FacilityID',columns = 'FlowName', values = 'FlowAmount').reset_index()
        
        
        #Merging dataframes together
        egrid = egrid1.merge(egrid3, left_on = 'FacilityID', right_on = 'FacilityID')
        egrid[['Heat input','Net generation']] = egrid[['Heat input','Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        
        
        #Calculating efficiency
        egrid.loc[:,"Efficiency"] = egrid['Net generation'].values*100/egrid['Heat input'].values
        egrid = egrid.sort_values(by = ['Efficiency'], axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last')
       
        #Replace inf
        egrid = egrid.replace([np.inf, -np.inf], np.nan)
        egrid = egrid.dropna(subset = ['Efficiency'])
       
        #Dropping less than 10%
        egrid_new1 = egrid[(egrid['Efficiency'] >= a) & (egrid['Efficiency'] <= b)]
        
        egrid_new1.to_csv('chk.csv')
        
        #Going back to per MJ units
        #egrid_new1['Carbon dioxide'] = egrid_new1['Carbon dioxide']/egrid_new1['Net generation']
        #egrid_new1['Sulfur dioxide'] = egrid_new1['Sulfur dioxide']/egrid_new1['Net generation']
        #egrid_new1['Methane'] = egrid_new1['Methane']/egrid_new1['Net generation']
        #egrid_new1['Nitrogen oxides'] = egrid_new1['Nitrogen oxides']/egrid_new1['Net generation']
        #egrid_new1['Nitrous oxide'] =  egrid_new1['Nitrous oxide']/egrid_new1['Net generation']
        
        
        
        

        
        col = list(egrid_new1['Heat input'])        
        egrid_new1['HeatInput'] = col;
        cols = egrid_new1.columns.tolist()
        egrid_new1 = egrid_new1[[cols[-1]] + cols[:-1]] 
        egrid_new1 = egrid_new1.drop(['Heat input','Efficiency'],axis = 1)
        
        
        return egrid_new1


def tri_func():
        
        #READING tri database
        tri = pd.read_csv("TRI_2016.csv", header=0, error_bad_lines=False)  
        
        tri = tri.drop_duplicates(keep = 'first')
        tri = tri[['FacilityID','FlowName','Compartment','FlowAmount']]
        tri2 = tri.groupby(['FacilityID','FlowName','Compartment'])['FlowAmount'].sum()
        tri2 = tri2.reset_index()
        tri2 = tri2.dropna(axis = 0, how = 'any')
        #TRI reshaping pivot not working so using pivot table
        tri3 = pd.pivot_table(tri2,index = ['FacilityID','Compartment'], columns = 'FlowName', values = 'FlowAmount')
        
        tri3 = tri3.reset_index()
        #print(tri2[['ReleaseType']])
        
        tri2egrid = pd.read_csv("tri_egrid.csv", usecols = ['EGRID_ID','TRI_ID1'],header=0, error_bad_lines=False)
        
        #Merging with the bridge file with EGRID
        tri4 = tri2egrid.merge(tri3, left_on = 'TRI_ID1', right_on = 'FacilityID')
        tri4 = tri4.drop(columns = ['FacilityID','TRI_ID1'])
        return tri4
        


def merge(a,b):
        
        #Merging Egrid and TRI based on a left biased merging so as not to lose egrid facilities
        egrid = egrid_func(a,b)
        
        tri = tri_func()
        
        database = egrid.merge(tri, how = 'left',left_on ='FacilityID', right_on = 'EGRID_ID')
        
        database = database.drop(columns = ['EGRID_ID'])
        database = database.dropna(axis = 1, how = 'all')
        database[['Compartment']] = database[['Compartment']].fillna(value = 'air')
        
        
        
        database[['Net generation']] = database[['Net generation']].apply(pd.to_numeric,errors = 'coerce')
        
        #for i in range(17,len(database.axes[1])):
        # database.iloc[:,i] = database.iloc[:,i]/database.iloc[:,12]
        
        col = list(database['Net generation'])         
        database['NetGen'] = col;
        cols = database.columns.tolist()
        database = database[[cols[-1]] + cols[:-1]] 
        database = database.drop(columns = ['Net generation'])
        #database.to_csv('chk.csv')
        return database



#This function is necessary to creates new line in the template for writing input and output flows. 
def createblnkrow(br,io):
    
    Column_number = 43 #THis is based on the template number of columns in the template. 
    for i in range(1,Column_number):
      
      v = io.cell(row = br,column = i).value
      io.cell(row = br+1 ,column = i).value = v
      io.cell(row = br+1 ,column = i).font = copy(io.cell(row = br,column = i).font)
      io.cell(row = br+1 ,column = i).border = copy(io.cell(row = br,column = i).border)
      io.cell(row = br+1 ,column = i).fill = copy(io.cell(row = br,column = i).fill)
    br = br + 1;
      #print(io.cell(row = 6,column = 4).value);
    return br



def generator(a,b,Reg):
    
            database = merge(a,b)
            
            #Reg = 'AZNM'
            #Reg = input('Please enter 4 lettered EGRID region here in uppercase: ')  
            database = database[database['eGRID subregion acronym'] == Reg]
            
                            
            
            fuel_name = pd.read_csv("fuelname.csv", header=0, error_bad_lines=False)
            for row in fuel_name.itertuples():
              #assuming the starting row to fill in the emissions is the 5th row Blank
              
              global blank_row;    
              blank_row = 6;
              index = 2;
              for roww in database.itertuples():
                
                
                #The fuels and their types are in two different columns
                if row[1] == roww[9] or row[1] == roww[10]:              
                    
                    fuelname = row[2]
                    fuelheat = row[3]
                    print(fuelname)
                                      
        
                    #Read the template files.                         
                    wbook = openpyxl.load_workbook('Fed_template_Main_19Dec2016.xlsx')
                    gi = wbook['General information']
                    
                    #This block is used for filling up the General Infomration Sheet      
                    gi['D4'].value = 'Electricity'
                    
                    gi['D5'].value = 'from '+ fuelname;
                    
                    gi['D6'].value = 'at generating facility';
                    
                    gi['D7'].value = 'Production Mix'; 
                    
                    
                    
                    gi['D10'].value = 'Electricity from '+ fuelname+' using power plants in the '+Reg+' region'
                    
                    #This function is used for the name of the fuels.
                    #def pm1(fn):
                    #    for row in primemover.itertuples():
                    #       if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                    #          if row[4] != None:                   
                    #             return row[4].capitalize()
                    
                    gi['D11'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+str(fuelname)
                    
                    gi['D12'].value = 'FALSE'
                    
                    gi['D14'].value = 'Electricity; voltage?'
                    
                    gi['D16'].value = '01/01/2017'
                    
                    gi['D17'].value = '12/31/2017'
                    
                    gi['D18'].value = '2014'
                    
                    gi['D20'].value = 'US-eGRID-'+Reg
                
                    v= [];
                    v1= [];
                    '''
                    #THis function is used for finding the NERC regions. THese input files are built manually. 
                    for row in temp.itertuples():
                        if row[2] == Reg:
                           v = row[1]
                           
                    for row in state.itertuples():
                        if row[1] == Reg:
                            v1 = row[2];
                    
                    gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '+v+'\nStates totally or partially included: '+v1
                    '''
                    
                    for rowww in database.itertuples():
                        if row[1] == roww[9] or row[1] == roww[10]:
                            v1.append(rowww[4])
                            v.append(roww[11])
                            
                    v2 = list(set(v1)) 
                    
                    str1 = ','.join(v2)
                    
                    v2 = list(set(v))
                    str2 = ','.join(v2)
                    gi['D24'].value = 'database subregion definitions:<https://www.epa.gov/energy/emissions-generation-resource-integrated-database-database>\nNERC region: '+str2+'\nStates totally or partially included: '+str1
                   
        
                    '''
                    #This is for printing the Prime Movers. Written to a output file and reread back. 
                    def pm(fn,rg):
                        f = open('out.txt', 'w')
                        for row in primemover.itertuples():
                          if row[2] == rg:
                            if row[3]  == fn or row[4] == fn or row[5] == fn or row[6] == fn:
                              if row[7] != None:                   
                                 f.write(row[7])
                                 f.write(',')
                        f.close()        
                    
                     
                    pm(Fuel,Reg)
                    
                    #linestring = open('out.txt', 'r').read()
                    
                    gi['D26'].value = 'This is an aggregation of technology types within this eGRID subregion.  List of prime movers:'+linestring
                    
                    '''
                    
               
                    #croppping the database according to the current fuel being considered
                    database_f1 = database[database['Plant primary coal/oil/gas/ other fossil fuel category'] == row[1]]
                    if database_f1.empty == True:
                          database_f1 = database[database['Plant primary fuel'] == row[1]]              
                    
                    
                  
                    #This block is used for filling up the InputOutput Sheet ONLY EMISSIONS
                    io = wbook['InputsOutputs']
                    
                    

                    #Fillin up with reference only only
                    io['A5'].value = 1
                    io['C5'].value =  0;
                    io['D5'].value =  'Electricity from '+fuelname;
                    io['E5'].value = '22: Utilities/2211: Electric Power Generation, Transmission and Distribution/'+str(fuelname)
                    io['G5'].value =  1 
                    io['H5'].value =  'MJ'
                    row1 = blank_row;
                    
                    blank_row = createblnkrow(blank_row,io)
                    
                    
                    if database_f1['HeatInput'].mean() != 0 and fuelheat != 0:
                    #Fillin up the fuel flow
                        io[row1][0].value = index;
                        io[row1][1].value = 5;
                        io[row1][3].value = fuelname;
                        io[row1][4].value = 'Input Fuel'
                        io[row1][6].value = np.sum(database_f1['HeatInput'])/np.sum(database_f1['NetGen']);
                        io[row1][7].value = 'kg';
                        io[row1][21].value = 'database data with plants over 10% efficiency';
                        io[row1][26].value = 'Normal';
                        io[row1][27].value = np.sum(database_f1['HeatInput'])/np.sum(database_f1['NetGen']);
                        io[row1][28].value = database_f1['HeatInput'].std();
                        io[row1][29].value = database_f1['HeatInput'].min();
                        io[row1][30].value = database_f1['HeatInput'].max();
                        row1 = blank_row;
                        index = index + 1;
                    
                    
                    
                    
                    
                    
                    #air 
                    d1 = database_f1[database_f1['Compartment']=='air']
                    d1 = d1.drop(columns = ['Compartment'])
                    d1.to_csv('chk.csv')
                    (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'air')
                    
                    #water
                    d1 = database_f1[database_f1['Compartment']=='water']
                    d1 = d1.drop(columns = ['Compartment'])
                    (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'water')
                    
                    #soil
                    d1 = database_f1[database_f1['Compartment']=='soil']
                    d1 = d1.drop(columns = ['Compartment'])
                    (blank_row,row1,index) = flowwriter(d1,io,row1,index,blank_row,'soil')
                    
                    

                            
                            
                            
                            
                            
                            
                            

                    filename = fuelname+"_"+Reg+".xlsx"
                    data_dir = os.path.dirname(os.path.realpath(__file__))+"\\results\\"
                    wbook.save(data_dir+filename)
                    break;


def flowwriter(database_f1,io,row1,index,blank_row,comp):
    
    flownames = list(database_f1.columns.values) 
    u = 17;             
    if comp == 'air':
        u = 11
        
    for i in range(u,len(flownames)):
    
        if np.isnan(database_f1[flownames[i]].mean()) == False:
            
            blank_row = createblnkrow(blank_row,io)
            io[row1][0].value = index;
            io[row1][2].value = 4;
            io[row1][3].value = flownames[i];
            io[row1][4].value = 'Elementary Flows/'+str(comp)+'/unspecified'
            io[row1][6].value = np.sum(database_f1[flownames[i]])/np.sum(database_f1['NetGen']);
            io[row1][7].value = 'kg';
            io[row1][21].value = 'database data with plants over 10% efficiency';
            io[row1][26].value = 'Normal';
            io[row1][27].value = np.sum(database_f1[flownames[i]])/np.sum(database_f1['NetGen']);
            io[row1][28].value = database_f1[flownames[i]].std();
            io[row1][29].value = database_f1[flownames[i]].min();
            io[row1][30].value = database_f1[flownames[i]].max();
            row1 = row1+1;
            index = index+1;
            
    return (blank_row,row1,index)





#p = egrid_func(10,100)
p = generator(10,100,'AZNM')
#p = merge(10,100)                  
#p = tri_func()